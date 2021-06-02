#!/usr/bin/env python3

"""
This module contains helper functions to work with xlsx templates
used in the service.
"""

import pandas as pd

from datetime import date
from io import BytesIO
from itertools import islice

from dateutil.relativedelta import relativedelta
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from typing import Collection


def get_template(calc_date: date,
                 indicator_names: Collection[str],
                 modeling_horizon: int) -> bytes:
    wb = Workbook()

    sheet = wb['Sheet']
    sheet.title = 'Values'

    # Set the worksheet header
    for row in sheet.iter_rows(max_row=1,
                               max_col=len(indicator_names)):
        for indicator_name, cell in zip(indicator_names, row):
            cell.value = indicator_name

    # Set date column
    rel_date = calc_date
    for col in sheet.iter_cols(max_col=1, min_row=2,
                               max_row=modeling_horizon + 1):
        for cell in col:
            rel_date = rel_date + relativedelta(months=3)
            cell.value = rel_date

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)

        return tmp.read()


def read_excel_bytes(in_bytes: bytes):
    with BytesIO(in_bytes) as f:
        wb = load_workbook(f)

    ws = wb.active
    df = ws.values

    cols = next(df)[1:]
    df = list(df)

    idx = tuple(r[0] for r in df)
    df = (islice(r, 1, None) for r in df)
    df = pd.DataFrame(df, index=idx, columns=cols).fillna(0)

    return df
